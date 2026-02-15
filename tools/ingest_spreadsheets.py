#!/usr/bin/env python3
"""
Ingest native spreadsheet files (XLSX, XLS, CSV) into full_text_corpus.db.

Each worksheet becomes a "page" entry with cells rendered as tab-separated
text, making the content searchable via the same FTS5 interface as PDFs.

PRIVACY NOTE: EFTA00029272 and EFTA00029480 contain victim pseudonym-to-name
mappings. The raw cell data is ingested for search purposes but victim names
must NEVER be published in reports or public exports.
"""

import csv
import os
import sqlite3
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    import xlrd
except ImportError:
    print("ERROR: xlrd not installed. Run: pip install xlrd")
    sys.exit(1)

DB_PATH = "/atb-data/rye/dump/epstein_files/full_text_corpus.db"
NATIVES_DIR = "/atb-data/rye/dump/epstein_files/datasets/dataset8/VOL00008/NATIVES/0001"

# All spreadsheet native files in DS8
SPREADSHEETS = [
    ("EFTA00015032", "xlsx"),
    ("EFTA00016155", "xlsx"),
    ("EFTA00016338", "csv"),
    ("EFTA00016339", "csv"),
    ("EFTA00016340", "csv"),
    ("EFTA00016341", "csv"),
    ("EFTA00018595", "xlsx"),
    ("EFTA00019180", "xls"),
    ("EFTA00019181", "xls"),
    ("EFTA00020242", "xlsx"),
    ("EFTA00023539", "xlsx"),
    ("EFTA00024809", "xlsx"),
    ("EFTA00024813", "xlsx"),
    ("EFTA00029272", "xlsx"),
    ("EFTA00029480", "xlsx"),
]


def read_xlsx(path):
    """Read an XLSX file, return list of (sheet_name, text_content) tuples."""
    sheets = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        lines = []
        for row in ws.iter_rows(values_only=True):
            cells = []
            for cell in row:
                if cell is not None:
                    cells.append(str(cell).strip())
                else:
                    cells.append("")
            # Skip entirely empty rows
            if any(c for c in cells):
                lines.append("\t".join(cells))
        text = "\n".join(lines)
        if text.strip():
            sheets.append((sname, text))
    wb.close()
    return sheets


def read_xls(path):
    """Read an XLS file, return list of (sheet_name, text_content) tuples."""
    sheets = []
    wb = xlrd.open_workbook(path)
    for sname in wb.sheet_names():
        ws = wb.sheet_by_name(sname)
        if ws.nrows == 0:
            continue
        lines = []
        for r in range(ws.nrows):
            cells = []
            for c in range(ws.ncols):
                val = ws.cell_value(r, c)
                if val is not None and str(val).strip():
                    cells.append(str(val).strip())
                else:
                    cells.append("")
            if any(c for c in cells):
                lines.append("\t".join(cells))
        text = "\n".join(lines)
        if text.strip():
            sheets.append((sname, text))
    return sheets


def read_csv_file(path):
    """Read a CSV file, return list of (sheet_name, text_content) tuples."""
    lines = []
    with open(path, 'r', errors='replace') as f:
        reader = csv.reader(f)
        for row in reader:
            cells = [c.strip() for c in row]
            if any(c for c in cells):
                lines.append("\t".join(cells))
    text = "\n".join(lines)
    return [("Sheet1", text)] if text.strip() else []


def update_document_record(conn, efta, file_path, num_sheets, file_size):
    """Update the existing document record to reflect native file ingestion."""
    cur = conn.cursor()

    # Check if document exists
    cur.execute("SELECT id, total_pages FROM documents WHERE efta_number = ?", (efta,))
    row = cur.fetchone()

    if row:
        doc_id = row[0]
        old_pages = row[1] or 0
        # Update to reflect native content (keep existing PDF pages, add native sheets)
        # We won't change total_pages since that refers to PDF pages
        # Instead we'll note the native file in the file_path
        return doc_id, old_pages
    else:
        # Create new document record
        cur.execute("""
            INSERT INTO documents (efta_number, dataset, file_path, total_pages, extraction_timestamp, file_size)
            VALUES (?, 8, ?, ?, datetime('now'), ?)
        """, (efta, file_path, num_sheets, file_size))
        conn.commit()
        return cur.lastrowid, 0


def ingest_spreadsheet(conn, efta, ext):
    """Ingest a single spreadsheet into full_text_corpus.db."""
    fname = f"{efta}.{ext}"
    path = os.path.join(NATIVES_DIR, fname)

    if not os.path.exists(path):
        print(f"  NOT FOUND: {path}")
        return False

    file_size = os.path.getsize(path)

    # Read the spreadsheet
    if ext == "xlsx":
        sheets = read_xlsx(path)
    elif ext == "xls":
        sheets = read_xls(path)
    elif ext == "csv":
        sheets = read_csv_file(path)
    else:
        print(f"  UNSUPPORTED: {ext}")
        return False

    if not sheets:
        print(f"  EMPTY: {fname}")
        return False

    cur = conn.cursor()

    # Check if native content already exists (look for pages with high page numbers
    # that would indicate native sheet pages, or check for exact text match)
    cur.execute("""
        SELECT COUNT(*) FROM pages
        WHERE efta_number = ? AND page_number > 1000
    """, (efta,))
    existing = cur.fetchone()[0]
    if existing > 0:
        print(f"  ALREADY INGESTED: {efta} (has {existing} native sheet pages)")
        return False

    # Use page numbers starting at 10001 to distinguish from PDF pages
    # (no PDF has 10000+ pages, so this is safe)
    base_page = 10001
    total_chars = 0
    total_rows_inserted = 0

    for i, (sheet_name, text) in enumerate(sheets):
        page_num = base_page + i
        # Prepend sheet name as header
        full_text = f"[Native {ext.upper()} — Sheet: {sheet_name}]\n\n{text}"
        char_count = len(full_text)
        total_chars += char_count

        cur.execute("""
            INSERT INTO pages (efta_number, page_number, text_content, char_count)
            VALUES (?, ?, ?, ?)
        """, (efta, page_num, full_text, char_count))
        total_rows_inserted += 1

    conn.commit()

    # Update FTS index
    for i, (sheet_name, text) in enumerate(sheets):
        page_num = base_page + i
        full_text = f"[Native {ext.upper()} — Sheet: {sheet_name}]\n\n{text}"
        # Get the rowid of the page we just inserted
        cur.execute("""
            SELECT id FROM pages
            WHERE efta_number = ? AND page_number = ?
        """, (efta, page_num))
        row = cur.fetchone()
        if row:
            try:
                cur.execute("""
                    INSERT INTO pages_fts(rowid, efta_number, page_number, text_content)
                    VALUES (?, ?, ?, ?)
                """, (row[0], efta, page_num, full_text))
            except Exception as e:
                # FTS insert may fail if entry already exists
                pass

    conn.commit()

    sheet_count = len(sheets)
    line_count = sum(text.count('\n') + 1 for _, text in sheets)
    print(f"  INGESTED: {fname} — {sheet_count} sheet(s), {line_count:,} rows, {total_chars:,} chars")
    return True


def main():
    print("=" * 60)
    print("INGESTING NATIVE SPREADSHEET FILES INTO full_text_corpus.db")
    print("=" * 60)
    print()

    conn = sqlite3.connect(DB_PATH)

    ingested = 0
    skipped = 0
    failed = 0

    for efta, ext in SPREADSHEETS:
        print(f"\n[{efta}.{ext}]")
        try:
            if ingest_spreadsheet(conn, efta, ext):
                ingested += 1
            else:
                skipped += 1
        except Exception as e:
            print(f"  ERROR: {e}")
            failed += 1

    conn.close()

    print()
    print("=" * 60)
    print(f"RESULTS: {ingested} ingested, {skipped} skipped, {failed} failed")
    print(f"Total: {len(SPREADSHEETS)} spreadsheet files")
    print("=" * 60)

    # Verify
    print("\nVerification:")
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM pages WHERE page_number >= 10001")
    native_pages = cur.fetchone()[0]
    cur.execute("SELECT SUM(char_count) FROM pages WHERE page_number >= 10001")
    native_chars = cur.fetchone()[0] or 0
    cur.execute("SELECT COUNT(DISTINCT efta_number) FROM pages WHERE page_number >= 10001")
    native_docs = cur.fetchone()[0]
    conn.close()

    print(f"  Native sheet pages in DB: {native_pages}")
    print(f"  Unique documents with native content: {native_docs}")
    print(f"  Total native text: {native_chars:,} characters")


if __name__ == '__main__':
    main()
